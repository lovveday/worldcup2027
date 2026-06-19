"""
World Cup 2026 Stats Tracker
Pulls fixtures, team stats, and player stats from API-Football and writes
them into data/world_cup_2026.xlsx.

Designed to be safe to run repeatedly (idempotent) and frugal with the
free-tier 100 requests/day quota: it only fetches stats for fixtures that
have finished AND that we haven't already stored.
"""

import os
import sys
import time
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API_KEY = os.environ.get("API_FOOTBALL_KEY")
if not API_KEY:
    print("ERROR: API_FOOTBALL_KEY environment variable not set.")
    sys.exit(1)

BASE_URL = "https://v3.football.api-sports.io"
HEADERS = {"x-apisports-key": API_KEY}

# World Cup 2026 league id in API-Football is 1, season 2026.
LEAGUE_ID = 1
SEASON = 2026

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
WORKBOOK_PATH = os.path.join(DATA_DIR, "world_cup_2026.xlsx")

REQUEST_DELAY = 1.1  # be polite to the free tier rate limit

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
NORMAL_FONT = Font(name="Arial", size=10)


def api_get(endpoint, params=None):
    """GET request against API-Football with basic error handling + rate-limit pause."""
    url = f"{BASE_URL}/{endpoint}"
    resp = requests.get(url, headers=HEADERS, params=params or {}, timeout=30)
    time.sleep(REQUEST_DELAY)
    if resp.status_code != 200:
        print(f"  WARNING: {endpoint} returned status {resp.status_code}: {resp.text[:200]}")
        return None
    payload = resp.json()
    remaining = resp.headers.get("x-ratelimit-requests-remaining")
    if remaining is not None:
        print(f"  (requests remaining today: {remaining})")
    if payload.get("errors"):
        print(f"  API errors on {endpoint}: {payload['errors']}")
    return payload.get("response", [])


def get_all_fixtures():
    print("Fetching full World Cup 2026 fixture list...")
    return api_get("fixtures", {"league": LEAGUE_ID, "season": SEASON}) or []


def get_fixture_stats(fixture_id):
    return api_get("fixtures/statistics", {"fixture": fixture_id}) or []


def get_fixture_players(fixture_id):
    return api_get("fixtures/players", {"fixture": fixture_id}) or []


def get_fixture_events(fixture_id):
    return api_get("fixtures/events", {"fixture": fixture_id}) or []


def ensure_workbook():
    """Create the workbook with all expected sheets/headers if it doesn't exist yet."""
    if os.path.exists(WORKBOOK_PATH):
        return load_workbook(WORKBOOK_PATH)

    wb = Workbook()

    matches = wb.active
    matches.title = "Matches"
    matches.append([
        "Fixture ID", "Date (UTC)", "Stage", "Group", "Venue", "City",
        "Home Team", "Away Team", "Home Score", "Away Score", "Status",
    ])

    team_stats = wb.create_sheet("Team Stats")
    team_stats.append([
        "Fixture ID", "Date (UTC)", "Team", "Opponent",
        "Shots on Goal", "Shots off Goal", "Total Shots", "Blocked Shots",
        "Corner Kicks", "Offsides", "Possession %", "Yellow Cards", "Red Cards",
        "Goalkeeper Saves", "Total Passes", "Passes Accurate", "Pass Accuracy %",
        "Fouls", "Expected Goals (xG)",
    ])

    player_stats = wb.create_sheet("Player Stats")
    player_stats.append([
        "Fixture ID", "Date (UTC)", "Team", "Player", "Position", "Minutes Played",
        "Rating", "Goals", "Assists", "Shots Total", "Shots on Target",
        "Passes Total", "Pass Accuracy %", "Tackles", "Interceptions",
        "Duels Won", "Duels Total", "Dribbles Won", "Dribbles Attempted",
        "Fouls Drawn", "Fouls Committed", "Yellow Cards", "Red Cards", "Saves",
    ])

    events = wb.create_sheet("Match Events")
    events.append([
        "Fixture ID", "Date (UTC)", "Minute", "Team", "Player", "Assist",
        "Type", "Detail", "Comments",
    ])

    log = wb.create_sheet("Update Log")
    log.append(["Run Timestamp (UTC)", "Fixtures Checked", "New Finished Matches Processed", "Notes"])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"

    return wb


def existing_fixture_ids(wb):
    sheet = wb["Matches"]
    return {row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value is not None}


def stats_already_collected(wb, fixture_id):
    sheet = wb["Team Stats"]
    return any(row[0].value == fixture_id for row in sheet.iter_rows(min_row=2))


def upsert_match_row(wb, fx):
    sheet = wb["Matches"]
    fixture_id = fx["fixture"]["id"]
    target_row = None
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == fixture_id:
            target_row = row[0].row
            break

    values = [
        fixture_id,
        fx["fixture"]["date"],
        fx["league"].get("round", ""),
        fx["league"].get("round", "") if "Group" in (fx["league"].get("round") or "") else "",
        (fx["fixture"].get("venue") or {}).get("name", ""),
        (fx["fixture"].get("venue") or {}).get("city", ""),
        fx["teams"]["home"]["name"],
        fx["teams"]["away"]["name"],
        fx["goals"]["home"],
        fx["goals"]["away"],
        fx["fixture"]["status"]["long"],
    ]

    if target_row:
        for col, val in enumerate(values, start=1):
            sheet.cell(row=target_row, column=col, value=val)
    else:
        sheet.append(values)


def add_team_stats(wb, fixture_id, date_str, stats_response):
    sheet = wb["Team Stats"]
    for team_block in stats_response:
        team_name = team_block["team"]["name"]
        opponent = None  # filled below by caller context if needed
        stat_map = {s["type"]: s["value"] for s in team_block.get("statistics", [])}

        def pct(val):
            if val is None:
                return None
            if isinstance(val, str) and val.endswith("%"):
                return val
            return val

        sheet.append([
            fixture_id, date_str, team_name, opponent,
            stat_map.get("Shots on Goal"),
            stat_map.get("Shots off Goal"),
            stat_map.get("Total Shots"),
            stat_map.get("Blocked Shots"),
            stat_map.get("Corner Kicks"),
            stat_map.get("Offsides"),
            pct(stat_map.get("Ball Possession")),
            stat_map.get("Yellow Cards"),
            stat_map.get("Red Cards"),
            stat_map.get("Goalkeeper Saves"),
            stat_map.get("Total passes"),
            stat_map.get("Passes accurate"),
            pct(stat_map.get("Passes %")),
            stat_map.get("Fouls"),
            stat_map.get("expected_goals"),
        ])


def add_player_stats(wb, fixture_id, date_str, players_response):
    sheet = wb["Player Stats"]
    for team_block in players_response:
        team_name = team_block["team"]["name"]
        for p in team_block.get("players", []):
            info = p["player"]
            stats = p["statistics"][0] if p.get("statistics") else {}
            games = stats.get("games", {}) or {}
            shots = stats.get("shots", {}) or {}
            passes = stats.get("passes", {}) or {}
            tackles = stats.get("tackles", {}) or {}
            duels = stats.get("duels", {}) or {}
            dribbles = stats.get("dribbles", {}) or {}
            fouls = stats.get("fouls", {}) or {}
            cards = stats.get("cards", {}) or {}
            goals = stats.get("goals", {}) or {}

            sheet.append([
                fixture_id, date_str, team_name, info.get("name"),
                games.get("position"), games.get("minutes"), games.get("rating"),
                goals.get("total"), goals.get("assists"),
                shots.get("total"), shots.get("on"),
                passes.get("total"), passes.get("accuracy"),
                tackles.get("total"), tackles.get("interceptions"),
                duels.get("won"), duels.get("total"),
                dribbles.get("success"), dribbles.get("attempts"),
                fouls.get("drawn"), fouls.get("committed"),
                cards.get("yellow"), cards.get("red"),
                goals.get("saves"),
            ])


def add_events(wb, fixture_id, date_str, events_response):
    sheet = wb["Match Events"]
    for ev in events_response:
        sheet.append([
            fixture_id, date_str, ev.get("time", {}).get("elapsed"),
            ev.get("team", {}).get("name"),
            (ev.get("player") or {}).get("name"),
            (ev.get("assist") or {}).get("name"),
            ev.get("type"), ev.get("detail"), ev.get("comments"),
        ])


def autosize_columns(wb):
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            sheet.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = ensure_workbook()

    fixtures = get_all_fixtures()
    print(f"Found {len(fixtures)} total fixtures in the tournament.")

    new_finished = 0
    for fx in fixtures:
        upsert_match_row(wb, fx)

        status_short = fx["fixture"]["status"]["short"]
        fixture_id = fx["fixture"]["id"]
        finished_statuses = {"FT", "AET", "PEN"}

        if status_short in finished_statuses and not stats_already_collected(wb, fixture_id):
            print(f"Processing finished match: {fx['teams']['home']['name']} vs {fx['teams']['away']['name']}")
            date_str = fx["fixture"]["date"]

            team_stats = get_fixture_stats(fixture_id)
            if team_stats:
                add_team_stats(wb, fixture_id, date_str, team_stats)

            player_stats = get_fixture_players(fixture_id)
            if player_stats:
                add_player_stats(wb, fixture_id, date_str, player_stats)

            events = get_fixture_events(fixture_id)
            if events:
                add_events(wb, fixture_id, date_str, events)

            new_finished += 1

    autosize_columns(wb)

    from datetime import datetime, timezone
    log_sheet = wb["Update Log"]
    log_sheet.append([
        datetime.now(timezone.utc).isoformat(timespec="seconds"),
        len(fixtures),
        new_finished,
        "OK",
    ])

    wb.save(WORKBOOK_PATH)
    print(f"Saved workbook to {WORKBOOK_PATH}. New finished matches processed: {new_finished}")


if __name__ == "__main__":
    main()
